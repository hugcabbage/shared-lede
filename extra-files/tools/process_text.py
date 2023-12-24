"""Text processing"""
import json
import os
import re
import shutil
import sys


def mlength(seq) -> int:
    """Get the length of the longest element in the sequence"""

    lengths = []
    for item in seq:
        if isinstance(item, (list, tuple)):
            mitem = mlength(item)
            lengths.append(mitem)
        else:
            lengths.append(len(str(item)))

    try:
        return max(lengths)
    except ValueError:
        return 0


def rectangles(*lists) -> tuple[list, list]:
    """Get the length (number of elements) and 
    width (length of the longest element) of multiple lists
    """

    lengths = []
    widths = []
    for lst in lists:
        lengths.append(len(lst))
        widths.append(mlength(lst))
    return lengths, widths


def to_markdown_table(*lists) -> str:
    """Convert multiple lists to markdown tables"""

    # Calculate the maximum width of each column and determine whether all columns have the same length
    col_lengths, col_widths = rectangles(*lists)
    if len(set(col_lengths)) != 1:
        print('The list length is not uniform')
        sys.exit()

    # Create table header
    header = '|' + '|'.join(str(lst[0]).center(col_widths[i])
                            for i, lst in enumerate(lists)) + '|\n'
    separator = '|' + \
                '|'.join('-' * (col_widths[i]) for i in range(len(lists))) + '|\n'

    # Create table content
    rows = ''
    for i in range(1, col_lengths[0]):
        row = '|'
        for j, lst in enumerate(lists):
            if i < len(lst):
                if isinstance(lst[i], (list, tuple)):
                    row += '<br>'.join(map(str, lst[i])).center(col_widths[j])
                else:
                    row += str(lst[i]).center(col_widths[j])
            row += '|'
        row += '\n'
        rows += row

    # Return the complete Markdown table
    return header + separator + rows


def manifest_to_lists(file) -> tuple[list, list]:
    """Convert manifest file into two lists"""

    if not os.path.isfile(file):
        print(f'Error: {file} does not exist.')
        sys.exit()

    with open(file, encoding='utf-8') as f:
        lines = f.readlines()

    package = []
    version = []
    for line in lines:
        parts = line.strip().split(' - ')
        if len(parts) != 2:
            print(f'Error: invalid line format -> {line}')
            continue
        if parts[0] == 'kernel':
            parts[1] = parts[1][:-26]
        if parts[0].startswith('kmod-'):
            continue
        if parts[0].startswith('lib'):
            continue
        if 'i18n' in parts[0]:
            continue
        if 'lib-' in parts[0]:
            continue
        if 'mod-' in parts[0]:
            continue
        package.append(parts[0])
        version.append(parts[1])

    return package, version


def xlsx_to_dict(file) -> dict:
    """convert xlsx file to dict"""
    from openpyxl import load_workbook

    wb = load_workbook(file)
    sheet = wb.active
    dict_data = {}

    # exclude title row, first column is key, convert other columns to list as value
    for i in range(2, sheet.max_row + 1):
        key = sheet.cell(row=i, column=1).value
        value = []
        for j in range(2, sheet.max_column + 1):
            value.append(sheet.cell(row=i, column=j).value)
        dict_data[key] = value

    return dict_data


def dict_to_json(data: dict) -> str:
    """convert dict data to json file, apply to header.json"""

    json_str = json.dumps(data, indent=2)

    # make key and value be one line
    json_str = json_str.replace('[\n    ', '[').replace(
        ',\n    ', ', ').replace('"\n  ', '"')
    return json_str


def get_remain_text(config) -> list:
    with open(config, encoding='utf-8') as f:
        text = f.readlines()

    s, e = 0, 0
    for i, line in enumerate(text):
        if all(x in line for x in ('CONFIG_TARGET_', 'DEVICE')):
            s = i + 1
        elif '# Applications' in line:
            e = i
            break
    while not text[s].strip():
        s += 1
    while not text[e - 1].strip():
        e -= 1

    return text[s:e]


def generate_header(headers: dict, model: str) -> list:
    t1, t2, t3 = headers[model][1:4]
    header = [
        f'CONFIG_TARGET_{t1}=y\n',
        f'CONFIG_TARGET_{t1}_{t2}=y\n',
        f'CONFIG_TARGET_{t1}_{t2}_DEVICE_{t3}=y\n'
    ]
    return header


def get_header_index(config) -> list:
    h_index = []
    with open(config, encoding='utf-8') as f:
        for i, line in enumerate(f):
            if line.startswith('CONFIG_TARGET') and line.endswith('=y\n'):
                h_index.append(i)
                if len(h_index) == 3:
                    break
            if i > 500:
                break
    return h_index


def modify_config_header(config, header: list, new_file=None):
    """Replace or add header"""

    h_index = get_header_index(config)
    with open(config, 'r+', encoding='utf-8') as f:
        content = f.readlines()
        if h_index:
            for i in range(3):
                content[h_index[i]] = header[i]
        else:
            content[0:0] = header
        if not new_file:
            f.seek(0, 0)
            f.writelines(content)
            return
    with open(new_file, 'w', encoding='utf-8') as f:
        f.writelines(content)


def simplify_config(config, *, backup=True, remain_text: list = None, keep_header=False):
    """Simplify .config file, keep only applications and themes"""

    if backup:
        shutil.copyfile(config, config + '.fullbak')

    header, apps, themes = [], [], []
    with open(config, encoding='utf-8') as f:
        apps_range, themes_range = False, False
        for line in f:
            if len(header) < 3 and line.startswith('CONFIG_TARGET') and line.endswith('=y\n'):
                header.append(line)
            elif re.match(r'# \d+\. Applications', line):
                apps_range = True
            elif re.match(r'# end of \d+\. Applications', line):
                apps_range = False
            elif apps_range:
                apps.append(line)
            elif re.match(r'# \d+\. Themes', line):
                themes_range = True
            elif re.match(r'# end of \d+\. Themes', line):
                themes_range = False
            elif themes_range:
                themes.append(line)

    if keep_header:
        header.insert(0, '# Target\n')
        header.append('\n')
    else:
        header.clear()

    if remain_text:
        remain_text.append('\n')
    else:
        remain_text = []

    re_remove_text = [
        r'^\n$',
        r'^#\n$',
        r'^# Configuration',
        r'^# end of'
    ]

    apps = ['# Applications\n'] + [x for x in apps if not any(re.match(r, x) for r in re_remove_text)] + ['\n']
    themes = ['# Themes\n'] + [x for x in themes if not any(re.match(r, x) for r in re_remove_text)]

    text = header + remain_text + apps + themes

    with open(config, 'w', encoding='utf-8') as f:
        f.writelines(text)
