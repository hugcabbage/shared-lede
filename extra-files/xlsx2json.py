# convert xlsx file to dict
def xlsx_to_dict(file_name):
    from openpyxl import load_workbook
    wb = load_workbook(file_name)
    sheet = wb.active
    dict = {}
    # exclude title row
    # first column is key, convert other columns to list as value
    for i in range(2, sheet.max_row + 1):
        key = sheet.cell(row=i, column=1).value
        value = []
        for j in range(2, sheet.max_column + 1):
            value.append(sheet.cell(row=i, column=j).value)
        dict[key] = value
    return dict


# convert dict to json file
def dict_to_json(dict, file_name):
    import json
    json_str = json.dumps(dict, indent=2)
    # make key and value be one line
    json_str = json_str.replace('[\n    ', '[').replace(',\n    ', ', ').replace('"\n  ', '"')
    with open(file_name, 'w') as f:
        f.write(json_str)
    print('json file saved')


if __name__ == '__main__':
    dict_to_json(xlsx_to_dict('headers.xlsx'), 'headers.json')
